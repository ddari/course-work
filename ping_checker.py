import subprocess
import platform
import argparse
import re
import openpyxl


def ping_ip(ip_address: str, count=1) -> (bool, str):
    """
    Ping IP address and return tuple:
    On success:
        * True
        * ms : str
    On failure:
        * False
        * ''
    """
    # для разных систем разные запросы
    if platform.system().lower() == "windows":
        reply = subprocess.run(['ping', '-n ', str(count), ip_address],
                               stdout=subprocess.PIPE,
                               stderr=subprocess.PIPE,
                               encoding='cp866')
    else:
        reply = subprocess.run(['ping', '-c', str(count), '-n', ip_address],
                               stdout=subprocess.PIPE,
                               stderr=subprocess.PIPE,
                               encoding='utf-8')
    # найти ms
    pink_ms = re.search(r'=\d*[ms,мс]', reply.stdout)

    if pink_ms is not None:
        return True, reply.stdout[pink_ms.start() + 1:pink_ms.end() - 1]
    else:
        return False, ''


def list_from_exel(file: str, paper: [str, None]) -> list:
    """
    From the file, where first column is IPs make list of ip
    """
    if paper is None:
        paper = 'Лист1'
    l = []
    k = 2  # начальный столбец

    # открытие для чтения
    wb = openpyxl.load_workbook(filename=file+'.xlsx', read_only=True)
    sheet = wb[paper]
    a = sheet.cell(row=k, column=1).value
    # пока не дойдёт до конца спика в файле добавляет значения
    while a is not None:
        l.append(a)
        k += 1
        a = sheet.cell(row=k, column=1).value
    return l


def list_to_ping_list(l: [list,tuple]) -> [(str, str)]:
    """
    From the list of IP make list of tuple,
    Which is result of ping and
    On success:
        * IP
        * ms
    On failure:
        * IP
        * 'Ошибка соединения'
    """
    ll = []
    for i in l:
        t = ping_ip(i)
        if t[0]:
            ll.append((i, t[1]))
        else:
            ll.append((i, 'Ошибка соединения'))
    return ll


def print_single_ip(ip):
    """
    Print result of single ping
    """
    go, ms = ping_ip(ip, 1)
    if go:
        print(f'Ping {ms} мс')
    else:
        print('Ответ не вернулся')


def print_many_ip(file, sheet='Лист1'):
    """
    Print result of many pings
    """
    ip_list = list_from_exel(file, sheet)
    ping_list = list_to_ping_list(ip_list)
    for i in range(len(ping_list)):
        print(f'{i + 1}. {ping_list[i][0]} {ping_list[i][1]}')


def give_file(fileIn, fileOut, sheet='Лист1'):
    """
    Construct xlsx file where
    First column is IP
    Second is result of program work
    """
    ip_list = list_from_exel(fileIn, sheet)
    ping_list = list_to_ping_list(ip_list)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Лист1'
    sheet.cell(row=1, column=1).value = 'IP'
    sheet.cell(row=1, column=2).value = 'Ping'
    for i in range(len(ping_list)):
        sheet.cell(row=i + 2, column=1).value = ping_list[i][0]
        sheet.cell(row=i + 2, column=2).value = ping_list[i][1]
    print('Документ успешно сформирован')
    wb.save(fileOut + '.xlsx')


def repeat(file):
    """
    Ping IP from first column
    And add column of results to the end
    """
    ip_list = list_from_exel(file, None)
    ping_list = list_to_ping_list(ip_list)
    wb = openpyxl.load_workbook(filename=file + '.xlsx')
    sheet = wb['Лист1']
    k = 2
    a = sheet.cell(row=k, column=2).value
    # перебираем пока не конец строк
    while a is not None:
        column = 2
        b = sheet.cell(row=k, column=column).value
        # перебираем пока не конец столбцов
        # Оптимальнее сделать это 1 раз, а не для каждой строчки
        while b is not None:
            column += 1
            b = sheet.cell(row=k, column=column).value
        sheet.cell(row=k, column=column).value = ping_list[k - 2][1]
        k += 1
        a = sheet.cell(row=k, column=1).value
    wb.save(file + '.xlsx')


def find_dif(file):
    """
    Add to file two columns
    One is result of work ping program
    Another is result of average of all pings
    If one of results is not number average result will be 'Ответ не вернулся'
    """

    def sred_arf(lt: list):
        sumar = 0
        try:
            for i in lt:
                sumar += int(i)
        except:
            return 'Ответ не вернулся'
        return float(sumar) / len(lt)

    ip_list = list_from_exel(file, None)
    ping_list = list_to_ping_list(ip_list)
    wb = openpyxl.load_workbook(filename=file + '.xlsx')
    sheet = wb['Лист1']
    k = 2
    column = 3
    a = sheet.cell(row=k, column=2).value
    while a is not None:
        column = 2
        l = []
        b = sheet.cell(row=k, column=column).value
        while b is not None:
            try:
                l.append(int(b))
            except:
                l.append(b)
            column += 1
            b = sheet.cell(row=k, column=column).value
        l.append(ping_list[k - 2][1])
        sheet.cell(row=k, column=column + 1).value = sred_arf(l)
        sheet.cell(row=k, column=column).value = ping_list[k - 2][1]
        k += 1
        a = sheet.cell(row=k, column=1).value
    sheet.cell(row=1, column=column + 1).value = "Среднее арифметическое"
    wb.save(file + '.xlsx')


if __name__ == '__main__':
    # для запуска из командной строки с аргументами
    parser = argparse.ArgumentParser(description='Ping script')
    # одиночная проверка IP
    parser.add_argument('-a', action="store", dest="ip", type=str)  # '140.82.121.4'
    # кол-во раз для одиночной проверки
    parser.add_argument('-c', action="store", dest="count", default=1, type=int)
    # файл из которого считывать информацию с 2 строчки, поддерживает csv и xlsx
    parser.add_argument('-f', action="store", dest="file", type=str)
    # лист файла для считывания
    parser.add_argument('-l', action="store", dest="list", type=str)
    # вывод в файл после считывания файла
    parser.add_argument('-o', action="store", dest="out", type=str)
    # добавить столбец
    parser.add_argument('-r', action="store", dest="repeat", type=str)
    # сравнить все прошлые столбцы, добавить 1 и добавить среднее арифметическое
    parser.add_argument('-d', action="store", dest="dif", type=str)
    args = parser.parse_args()

    # проверка на наличие различных параметров
    if args.ip:
        print_single_ip(args.ip)
    if args.file:
        print_many_ip(args.file, args.list)
        if args.out:
            give_file(args.file, args.out, args.list)
    if args.repeat:
        try:
            repeat(args.repeat)
        except:
            print('файл не найден или его формат неверный')
    if args.dif:
        try:
            find_dif(args.dif)
        except:
            print('файл не найден или его формат неверный')
    # python ping_cheker.py -a 8.8.8.8 -f pings.xlsx -o out -r out -d out
